import os

from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from matrix import Matrix
import re

class TypeExtract:
    def __init__(self, add:"workbook", sheet:"worksheet"):
        self.__data = []
        self.__showall = True
        self.append_to_wb = None
        self.append_to_wb_obj = None
        self.final_wb = None
        self.final_sheet = None
        self.current_cell = None
        self.min_row_src = 0
        self.min_col_src = 0
        self.max_row_src = 0
        self.max_col_src = 0
        self.min_row_dst = 0
        self.min_col_dst = 0
        self.index = 0
        self.copy_data_signature = {}
        self.add = add
        self.new_wb = Workbook()
        try:
            self.wb = load_workbook(add)
        except Exception as e:
            print(e)
        try:
            if type(sheet) == str:
                self.sheet = self.wb[sheet]
            elif type(sheet) == int:
                self.sheet = self.wb.worksheets[sheet]
        except Exception as e:
            print(e)

        self.__header = None

    def __extract(self, personal_data=None, custom_condition:"code"=None, type_:"DataType"=None, min_row=None, min_col=None, max_row=None, max_col=None, use_header=True, first_row_header=True):

        if (max_row and max_col) and not (min_row and min_col):
            min_row = self.sheet.min_row
            min_col = self.sheet.min_column

        elif (min_row and min_col) and not (max_row and max_col):
            max_row = self.sheet.max_row
            max_col = self.sheet.max_column

        elif not (min_row and min_col and max_row and max_col):
            min_row = self.sheet.min_row
            max_row = self.sheet.max_row
            min_col = self.sheet.min_column
            max_col = self.sheet.max_column

        if not first_row_header:
            use_header = True

        if personal_data:
            transposed_data = Matrix(personal_data).transpose
            self.__header = {}
            for data in transposed_data:
                key = data[0]
                value = data[1:]
                self.__header[key] = value

            full_data = self.__header.copy()
            data_dict = full_data.copy()
        else:
            self.__header = dict.fromkeys(list(self.sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row, values_only=True))[0], False)
            full_data = self.__header.copy()
            for column, key in zip(self.sheet.iter_cols(min_col=min_col, min_row=min_row+1, max_col=max_col, max_row=max_row), full_data):
                full_data[key] = [item.value for item in column]

            remove_header = False
            skip_first = True
            index_for_rem_header = 0
            data_dict = dict.fromkeys(list(self.sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=min_row, values_only=True))[0], [])
            for row in self.sheet.iter_rows(min_row=min_row, min_col=min_col,
                                            max_row=max_row, max_col=max_col):
                for cell in row:
                    value = cell.value
                    coord = cell.coordinate
                    row_coord, col_coord = coordinate_from_string(coord)[1], column_index_from_string(coordinate_from_string(coord)[0])
                    header_val = self.sheet.cell(min_row, col_coord).value
                    column_val = self.sheet.cell(min_row, col_coord).value

                    if custom_condition:
                        if self.__custom_cond(custom_condition, value, column=column_val, data = full_data):
                            if min_row != row_coord:
                                self.__header[header_val] += 1

                                if data_dict[header_val] != None:
                                    col_val = data_dict[header_val].copy()
                                    col_val.append(value)
                                    data_dict[header_val] = col_val


                        elif remove_header:
                            try:
                                new_header_dict = {}
                                for k,v in self.__header.items():
                                    if self.__header[k] != 0:
                                        new_header_dict[k] = v

                                self.__header = new_header_dict

                            except KeyError:
                                remove_header = False

                        if not self.__custom_cond(custom_condition, value, column=column_val, data=full_data):
                            null_val = data_dict[header_val].copy()
                            null_val.append(None)
                            data_dict[header_val] = null_val

                    elif type_:
                        if type(value) == type_ and not first_row_header and not skip_first:
                            if min_row != row_coord:
                                self.__header[header_val] += 1
                                if data_dict[header_val] != None:
                                    col_val = data_dict[header_val].copy()
                                    col_val.append(value)
                                    data_dict[header_val] = col_val

                        elif remove_header:
                            try:
                                new_header_dict = {}
                                for k,v in self.__header.items():
                                    if self.__header[k] != 0:
                                        new_header_dict[k] = v

                                self.__header = new_header_dict

                            except KeyError:
                                remove_header = False

                        if not type_:
                            null_val = data_dict[header_val].copy()
                            null_val.append(None)
                            data_dict[header_val] = null_val

                    else:
                        if min_row != row_coord:
                            self.__header[header_val] += 1
                            if data_dict[header_val] != None:
                                col_val = data_dict[header_val].copy()
                                col_val.append(value)
                                data_dict[header_val] = col_val

                        elif remove_header:
                            try:
                                new_header_dict = {}
                                for k,v in self.__header.items():
                                    if self.__header[k] != 0:
                                        new_header_dict[k] = v

                                self.__header = new_header_dict

                            except KeyError:
                                remove_header = False


                if index_for_rem_header == max_row-2:
                    remove_header = True
                index_for_rem_header += 1

                first_row_header = False
                skip_first = False


        # Adding headers to the data
        v = 0
        for k in self.__header:
            try:
                self.__header[k] = data_dict[k]
            except IndexError:
                pass
            v += 1
        if use_header:
            self.__data.append(list(self.__header.keys()))

        final_data = Matrix(list(self.__header.values())).transpose

        for r in final_data:
            if self.__showall:
                if any(r):
                    self.__data.append(r)

            else:
                if all(r):
                    self.__data.append(r)

        if not self.__showall:
            new_data = Matrix(self.__data).transpose
            self.__header = {k:v[1:] for k,v in zip(list(self.__header.keys()), new_data)}



    def copy_data(self, personal_data=None, type_:"DataType"=None, custom_condition=None, min_row_source=None, min_col_source=None, max_row_source=None, max_col_source=None, use_header=True, first_row_header=True,
                  new_wb=None, copy_to_different_wb=False, sheet="Sheet", index=0, min_row_dst=1, min_col_dst=1):

        self.index = index
        self.min_row_src = min_row_source
        self.min_col_src = min_col_source
        self.max_row_src = max_row_source
        self.max_col_src = max_col_source
        self.min_row_dst = min_row_dst
        self.min_col_dst = min_col_dst

        if not sheet:
            sheet = "Sheet"

        self.copy_data_signature = dict(type_=type_, custom_condition=custom_condition, min_row_source=min_row_source, min_col_source=min_col_source,
                                        max_row_source=max_row_source, max_col_source=max_col_source, use_header=use_header, first_row_header=first_row_header,
                  new_wb=new_wb, copy_to_different_wb=copy_to_different_wb, sheet=sheet, index=index, min_row_dst=min_row_dst, min_col_dst=min_col_dst)

        # if not personal_data:
        self.__extract(personal_data=personal_data,custom_condition=custom_condition, type_=type_, min_row=min_row_source, min_col=min_col_source, max_row=max_row_source, max_col=max_col_source,
                            use_header=use_header, first_row_header=first_row_header)

        data = self.__data
        # for d in data:
        #     print(d)
        #
        # print("-----------------------------------------------------------------------------------------------------")

        if copy_to_different_wb:
            self.wb = load_workbook(copy_to_different_wb)
            self.add = copy_to_different_wb

            if not (sheet in self.wb):
                self.wb.create_sheet(sheet, index=index)
            new_sheet = self.wb[sheet]
            new_sheet.delete_cols(1, new_sheet.max_column)
            new_sheet.delete_rows(1, new_sheet.max_row)
            self.__custom_copy(new_sheet, data, min_row_dst, min_col_dst)
            self.final_wb = self.wb
            self.final_sheet = self.final_wb[sheet]

            return self

        elif new_wb and sheet:
            if not (sheet in self.new_wb):
                self.new_wb.create_sheet(sheet, index=index)
            new_sheet = self.new_wb[sheet]
            new_sheet.delete_cols(1, new_sheet.max_column)
            new_sheet.delete_rows(1, new_sheet.max_row)
            self.__custom_copy(new_sheet, data, min_row_dst, min_col_dst)
            self.final_wb = self.new_wb
            self.final_sheet = self.final_wb[sheet]

            return self

        elif not new_wb and sheet:
            if not (sheet in self.wb):
                self.wb.create_sheet(sheet, index=index)
            new_sheet = self.wb[sheet]
            new_sheet.delete_cols(1, new_sheet.max_column)
            new_sheet.delete_rows(1, new_sheet.max_row)
            self.__custom_copy(new_sheet, data, min_row_dst, min_col_dst)
            self.new_wb = self.wb
            self.final_wb = self.wb
            self.final_sheet = self.final_wb[sheet]

            return self


    def __custom_copy(self, new_sheet, data, min_row, min_col):
        max_row = len(data)
        max_col = len(data[0])
        r = 0
        for row in new_sheet.iter_rows(min_row=min_row, max_row= min_row+max_row-1, min_col=min_col, max_col = min_col + max_col-1):
            try:
                c = 0
                for cell in row:
                    try:
                        cell.value = data[r][c]
                    except IndexError:
                        continue
                    c += 1
            except IndexError:
                continue
            r += 1

    def append_to_rows(self, data, wb = None, sheet=None):
        if wb and sheet:
            if wb != self.append_to_wb:
                self.append_to_wb = wb
                self.append_to_wb_obj = load_workbook(wb)
                self.final_sheet = self.append_to_wb_obj[sheet]
            else:
                self.final_sheet = self.append_to_wb_obj[sheet]

            # wb = load_workbook(wb)
            # self.final_sheet = wb[sheet]

        elif sheet:
            self.final_sheet = self.final_wb[sheet]

        else:
            self.final_sheet = self.sheet

        for r in data:
            self.final_sheet.append(r)

        return self

    def append_to_cols(self, data, wb = None, sheet=None):
        if wb and sheet:
            if wb != self.append_to_wb:
                self.append_to_wb = wb
                self.append_to_wb_obj = load_workbook(wb)
                self.final_sheet = self.append_to_wb_obj[sheet]
            else:
                self.final_sheet = self.append_to_wb_obj[sheet]

        elif sheet:
            self.final_sheet = self.final_wb[sheet]

        else:
            self.final_sheet = self.sheet

        index_col = self.final_sheet.max_column+1
        index_row = self.final_sheet.min_row
        self.__custom_copy(new_sheet=self.final_sheet, data=data, min_row=index_row, min_col=index_col)


        return self

    def __custom_cond(self, code, value, column, data):

        locals()["showall"] = True
        initial_code = "condition = False\n"
        codes = code.split("\n")
        # initial_code.partition()
        for exp in codes:
            initial_code += f"{exp}\n"

        initial_code = initial_code.strip("\n").strip(" ")
        final_code = "condition = True"
        full_code = initial_code + final_code
        exec(full_code)
        self.__showall = locals()["showall"]
        return locals()["condition"]

    def execute(self, code, wb=None, sheet=None):
        if wb and sheet:
            if type(sheet) == str:
                current_sheet = wb[sheet]
            elif type(sheet) == int:
                current_sheet = wb[wb.worksheets[sheet]]

        elif sheet:
            if type(sheet) == str:
                current_sheet = self.final_wb[sheet]
            elif type(sheet) == int:
                current_sheet = self.final_wb[self.final_wb.worksheets[sheet]]

        else:
            current_sheet = self.final_sheet

        header = False
        if self.__data[0] == list(self.__header.keys()):
            header = self.__data[0]

        SHEET = current_sheet
        CELL = current_sheet.cell
        COLUMNS = self.__header

        if header:
            DATA = self.__data[1:]
        else:
            DATA = self.__data
        # DTYPE = {k:type(v[0]) for k,v in COLUMNS.items()}



        code = code.strip("\n")
        initial_code = ""
        codes = code.split("\n")
        for exp in codes:
            initial_code += f"{exp}\n"

        initial_code = initial_code.strip("\n").strip("\s")
        final_code = ""
        full_code = initial_code + final_code
        exec(full_code)

        self.__header = locals()["COLUMNS"]
        self.__data = Matrix(list(self.__header.values())).transpose

        if header:
            header = list(self.__header.keys())
            self.__data.insert(0, header)

        return self


    @property
    def appended_data(self):
        return self.final_sheet.values

    @property
    def get_data(self):
        return self.__data

    @property
    def get_header(self):
        return self.__header

    def save_(self, wb=None, dst=None, sheet=None):
        # if not (wb) and not (dst):
        #     wb = wb.final_wb
        #     wb.save(dst.get())

        if wb and dst:
            final_saved = self.copy_data(personal_data=self.get_data,type_=None, custom_condition=None, min_row_source=self.min_row_src, min_col_source=self.min_col_src,
                                        max_row_source=self.max_row_src, max_col_source=self.max_col_src, use_header=True, first_row_header=True,
                  new_wb=True, copy_to_different_wb=dst, sheet=sheet, index=self.index, min_row_dst=self.min_row_dst, min_col_dst=self.min_col_dst)

            final_saved.final_wb.save(dst)
            # final_saved.save_(wb=final_saved,dst=dst)

        else:
            final_saved = self.copy_data(personal_data=self.get_data, type_=None, custom_condition=None,
                                         min_row_source=self.min_row_src, min_col_source=self.min_col_src,
                                         max_row_source=self.max_row_src, max_col_source=self.max_col_src,
                                         use_header=True, first_row_header=True,
                                         new_wb=True, copy_to_different_wb=self.add, sheet=sheet, index=self.index,
                                         min_row_dst=self.min_row_dst, min_col_dst=self.min_col_dst)

            final_saved.final_wb.save(self.add)
            # final_saved.save_(wb = final_saved, dst=self.add)

        # if wb == self.new_wb and dst:
        #     try:
        #         wb.save(dst)
        #     except PermissionError:
        #         os.remove(dst)
        #         wb.save(dst)

        # elif wb == self.wb:
        #     try:
        #         wb.save(self.add)
        #     except PermissionError:
        #         os.remove(self.add)
        #         wb.save(self.add)
        #

